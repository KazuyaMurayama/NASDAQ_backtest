"""
Generate Excel with yearly returns + in-cell data bars
Negative bars extend left (red), positive bars extend right (blue)
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import DataBarRule
from copy import copy

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
yr_csv = os.path.join(BASE, 'yearly_returns_7strategies.csv')
yr = pd.read_csv(yr_csv, index_col=0)

order = ['DH Static (35/30/35)', 'DH Dynamic CAGR25+', 'A2 Optimized',
         'Ens2(Asym+Slope)', 'DD Only', 'BH 3x', 'BH 1x']
short = ['DH Static\n(35/30/35)', 'DH Dyn\nCAGR25+', 'A2\nOptimized',
         'Ens2\n(Asym+Slope)', 'DD\nOnly', 'BH\n3x', 'BH\n1x']
cagrs = [16.07, 25.23, 29.19, 22.20, 25.58, 19.21, 10.98]

wb = Workbook()
ws = wb.active
ws.title = "年次リターン"

# Styles
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=10, name='Arial')
nml_font = Font(size=10, name='Arial')
bld_font = Font(bold=True, size=10, name='Arial')
thin_bdr = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
yr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
cagr_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
sum_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Title
ws.merge_cells('A1:H1')
ws['A1'] = '7戦略 年次リターン（1974-2026）'
ws['A1'].font = Font(bold=True, size=14, name='Arial')

ws.merge_cells('A2:H2')
ws['A2'] = '条件: delay=2日, cost=0.86%/年 (TQQQ), データ=1974-2026 | * = 3資産ポートフォリオ (NASDAQ 3x + Gold + Bond)'
ws['A2'].font = Font(size=9, name='Arial', italic=True, color='666666')

# Row 3: descriptions
ws.cell(row=3, column=1, value='概要').font = bld_font
descs = ['A2+Gold30%\n+Bond35% *', 'A2動的配分\nCAGR25%+ *', 'DD+VT+Slope\n+MD+VIX', 'DD+AsymEWMA\n+Slope(旧)', 'DD制御のみ\n-18%/92%', 'NASDAQ 3倍\n無管理', 'NASDAQ指数\nレバなし']
for j, d in enumerate(descs):
    c = ws.cell(row=3, column=j+2, value=d)
    c.font = Font(size=8, name='Arial', color='666666')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[3].height = 30

# Row 4: CAGR
ws.cell(row=4, column=1, value='CAGR').font = bld_font
ws.cell(row=4, column=1).fill = cagr_fill
for j, c in enumerate(cagrs):
    cell = ws.cell(row=4, column=j+2, value=c/100)
    cell.number_format = '+0.00%;-0.00%'
    cell.font = Font(bold=True, size=10, name='Arial', color='2F5496')
    cell.fill = cagr_fill
    cell.alignment = Alignment(horizontal='center')

# Row 5: Headers
for j, sn in enumerate(['Year'] + list(short)):
    c = ws.cell(row=5, column=j+1, value=sn.replace('\n',' ') if j==0 else sn)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_bdr
ws.row_dimensions[5].height = 30

# Data rows (row 6+)
for i, year in enumerate(yr.index):
    row = 6 + i
    yc = ws.cell(row=row, column=1, value=year)
    yc.font = bld_font; yc.fill = yr_fill
    yc.alignment = Alignment(horizontal='center'); yc.border = thin_bdr

    for j, name in enumerate(order):
        val = yr.loc[year, name] / 100
        cell = ws.cell(row=row, column=j+2, value=val)
        cell.number_format = '+0.0%;-0.0%'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_bdr
        if val < 0:
            cell.font = Font(size=10, name='Arial', color='CC0000')
        else:
            cell.font = Font(size=10, name='Arial', color='2F5496')

last_data_row = 6 + len(yr.index) - 1

# Summary stats
summary_start = last_data_row + 2
stats = [
    ('中央値', lambda s: s.median()/100, True),
    ('最大', lambda s: s.max()/100, True),
    ('最小', lambda s: s.min()/100, True),
    ('標準偏差', lambda s: s.std()/100, True),
    ('プラス年数', lambda s: int((s>0).sum()), False),
    ('マイナス年数', lambda s: int((s<=0).sum()), False),
]
for si, (sn, func, is_pct) in enumerate(stats):
    r = summary_start + si
    sc = ws.cell(row=r, column=1, value=sn)
    sc.font = bld_font; sc.fill = sum_fill; sc.border = thin_bdr
    for j, name in enumerate(order):
        val = func(yr[name])
        cell = ws.cell(row=r, column=j+2, value=val)
        cell.number_format = '+0.0%;-0.0%' if is_pct else '0'
        cell.font = nml_font; cell.alignment = Alignment(horizontal='center')
        cell.border = thin_bdr; cell.fill = sum_fill

# Data bars: positive = blue right, negative = red left
# openpyxl DataBarRule with XML patching for negative color
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils.cell import get_column_letter

for j in range(len(order)):
    col = get_column_letter(j + 2)
    rng = f"{col}6:{col}{last_data_row}"

    rule = DataBarRule(
        start_type='num', start_value=-1.0,
        end_type='num', end_value=1.0,
        color='4472C4',
    )
    ws.conditional_formatting.add(rng, rule)

# Column widths
ws.column_dimensions['A'].width = 7
for j in range(len(order)):
    ws.column_dimensions[get_column_letter(j+2)].width = 16

ws.freeze_panes = 'B6'

# Save initial
out = os.path.join(BASE, 'YEARLY_RETURNS_7STRATEGIES.xlsx')
wb.save(out)

# Patch XML to add negative bar color (red)
import zipfile, shutil, tempfile, re

tmpdir = tempfile.mkdtemp()
tmp_xlsx = os.path.join(tmpdir, 'patched.xlsx')

with zipfile.ZipFile(out, 'r') as zin:
    with zipfile.ZipFile(tmp_xlsx, 'w') as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                content = data.decode('utf-8')
                # Add negativeBarColor after each dataBar closing cfvo
                # Pattern: find </dataBar> and insert negFillColor before it
                content = content.replace(
                    '</dataBar>',
                    '<negativeFillColor rgb="FFCC0000"/></dataBar>'
                )
                data = content.encode('utf-8')
            zout.writestr(item, data)

shutil.move(tmp_xlsx, out)
shutil.rmtree(tmpdir)

print(f"Saved: {out}")
print("Data bars: blue (positive, right) + red (negative, left)")
