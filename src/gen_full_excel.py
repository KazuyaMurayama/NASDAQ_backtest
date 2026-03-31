"""
Generate single-sheet Excel with ALL content from YEARLY_RETURNS_REPORT.md:
  - Conditions, strategy descriptions, summary stats
  - Yearly returns (1974-2026) with data bars
  - Monthly returns (2021-2026 OOS) with data bars
"""
import pandas as pd
import numpy as np
import os, zipfile, shutil, tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils.cell import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
yr = pd.read_csv(os.path.join(BASE, 'yearly_returns_7strategies.csv'), index_col=0)
mo = pd.read_csv(os.path.join(BASE, 'monthly_returns_oos.csv'), index_col=0)

order = ['DH Static (35/30/35)', 'DH Dynamic CAGR25+', 'A2 Optimized',
         'Ens2(Asym+Slope)', 'DD Only', 'BH 3x', 'BH 1x']
short = ['DH Static', 'DH Dyn 25+', 'A2 Opt', 'Ens2', 'DD Only', 'BH 3x', 'BH 1x']
cagrs = [16.07, 25.23, 29.19, 22.20, 25.58, 19.21, 10.98]
descs = [
    'A2+Gold30%+Bond35%\n四半期リバランス *',
    'A2動的配分\nCAGR25%+制約 *',
    'DD+AsymEWMA+Slope\n+MomDecel+VIX',
    'DD+AsymEWMA(20/5)\n+Slope(旧推奨)',
    'DD制御のみ\n-18%退避/92%復帰',
    'NASDAQ 3倍レバ\n無管理保有',
    'NASDAQ指数\nレバなし',
]

# Styles
HDR = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_F = Font(bold=True, color="FFFFFF", size=10, name='Arial')
SEC = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SEC_F = Font(bold=True, size=11, name='Arial', color="2F5496")
CAGR_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
YR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
COND_FILL = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
NF = Font(size=10, name='Arial')
BF = Font(bold=True, size=10, name='Arial')
BDR = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
)
CA = Alignment(horizontal='center', vertical='center')
LA = Alignment(horizontal='left', vertical='center')
WA = Alignment(horizontal='center', vertical='center', wrap_text=True)

def pos_neg_font(val):
    if val < 0:
        return Font(size=10, name='Arial', color='CC0000')
    return Font(size=10, name='Arial', color='2F5496')

def write_header_row(ws, row, col_start, labels):
    for j, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start+j, value=label)
        c.fill = HDR; c.font = HDR_F; c.alignment = WA; c.border = BDR

def add_data_bars(ws, col_letter, start_row, end_row):
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    rule = DataBarRule(start_type='num', start_value=-1.0,
                       end_type='num', end_value=1.0, color='4472C4')
    ws.conditional_formatting.add(rng, rule)

wb = Workbook()
ws = wb.active
ws.title = "年次・月次リターン"
r = 1  # current row

# ===== TITLE =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '7戦略 年次・月次リターン（1974-2026）'
ws[f'A{r}'].font = Font(bold=True, size=16, name='Arial')
r += 1

# ===== CONDITIONS =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '* = 3資産ポートフォリオ (NASDAQ 3x + Gold 447A + Bond 2621)'
ws[f'A{r}'].font = Font(size=9, name='Arial', italic=True, color='666666')
r += 2

# Section: 検証条件
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '■ 検証条件'
ws[f'A{r}'].font = SEC_F; ws[f'A{r}'].fill = SEC
r += 1

conds = [('データ期間', '1974-01-02 〜 2026-03-26'),
         ('実行遅延', '2営業日'), ('経費率', '年0.86%（TQQQ準拠）'),
         ('リバランス閾値', '20%'), ('OOS期間', '2021年5月〜（月次リターン部分）')]
for label, val in conds:
    ws.cell(row=r, column=1, value=label).font = BF
    ws.cell(row=r, column=1).fill = COND_FILL
    ws.cell(row=r, column=2, value=val).font = NF
    ws.cell(row=r, column=2).fill = COND_FILL
    ws.merge_cells(f'B{r}:D{r}')
    r += 1
r += 1

# ===== STRATEGY OVERVIEW =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '■ 各戦略の概要'
ws[f'A{r}'].font = SEC_F; ws[f'A{r}'].fill = SEC
r += 1

# Description row
ws.cell(row=r, column=1, value='概要').font = BF
for j, d in enumerate(descs):
    c = ws.cell(row=r, column=j+2, value=d)
    c.font = Font(size=8, name='Arial', color='666666')
    c.alignment = WA
ws.row_dimensions[r].height = 35
r += 1

# CAGR row
ws.cell(row=r, column=1, value='CAGR').font = BF
ws.cell(row=r, column=1).fill = CAGR_FILL
for j, cg in enumerate(cagrs):
    c = ws.cell(row=r, column=j+2, value=cg/100)
    c.number_format = '+0.00%;-0.00%'
    c.font = Font(bold=True, size=11, name='Arial', color='2F5496')
    c.fill = CAGR_FILL; c.alignment = CA
r += 1

# ===== SUMMARY STATS =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '■ 統計サマリー'
ws[f'A{r}'].font = SEC_F; ws[f'A{r}'].fill = SEC
r += 1

write_header_row(ws, r, 1, ['統計量'] + short)
r += 1

stats = [
    ('CAGR', lambda s,i: cagrs[i]/100, True),
    ('中央値', lambda s,i: s.median()/100, True),
    ('最大', lambda s,i: s.max()/100, True),
    ('最小', lambda s,i: s.min()/100, True),
    ('標準偏差', lambda s,i: s.std()/100, True),
    ('プラス年数', lambda s,i: int((s>0).sum()), False),
    ('マイナス年数', lambda s,i: int((s<=0).sum()), False),
]
for sn, func, is_pct in stats:
    ws.cell(row=r, column=1, value=sn).font = BF
    ws.cell(row=r, column=1).fill = SUM_FILL; ws.cell(row=r, column=1).border = BDR
    for j, name in enumerate(order):
        val = func(yr[name], j)
        c = ws.cell(row=r, column=j+2, value=val)
        c.number_format = '+0.0%;-0.0%' if is_pct else '0'
        c.font = NF; c.alignment = CA; c.fill = SUM_FILL; c.border = BDR
    r += 1
r += 1

# ===== YEARLY RETURNS =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '■ 年次リターン（1974-2026）'
ws[f'A{r}'].font = SEC_F; ws[f'A{r}'].fill = SEC
r += 1

write_header_row(ws, r, 1, ['Year'] + short)
r += 1
yr_data_start = r

for year in yr.index:
    yc = ws.cell(row=r, column=1, value=year)
    yc.font = BF; yc.fill = YR_FILL; yc.alignment = CA; yc.border = BDR
    for j, name in enumerate(order):
        val = yr.loc[year, name] / 100
        c = ws.cell(row=r, column=j+2, value=val)
        c.number_format = '+0.0%;-0.0%'
        c.font = pos_neg_font(val)
        c.alignment = CA; c.border = BDR
    r += 1

yr_data_end = r - 1

# Add data bars for yearly
for j in range(len(order)):
    add_data_bars(ws, get_column_letter(j+2), yr_data_start, yr_data_end)

r += 1

# ===== MONTHLY RETURNS (OOS) =====
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '■ 月次リターン（2021-2026, OOS期間）'
ws[f'A{r}'].font = SEC_F; ws[f'A{r}'].fill = SEC
r += 1

write_header_row(ws, r, 1, ['Year-Month'] + short)
r += 1
mo_data_start = r

for ym in mo.index:
    yc = ws.cell(row=r, column=1, value=str(ym))
    yc.font = BF; yc.fill = YR_FILL; yc.alignment = CA; yc.border = BDR
    for j, name in enumerate(order):
        val = mo.loc[ym, name] / 100
        c = ws.cell(row=r, column=j+2, value=val)
        c.number_format = '+0.0%;-0.0%'
        c.font = pos_neg_font(val)
        c.alignment = CA; c.border = BDR
    r += 1

mo_data_end = r - 1

# Add data bars for monthly
for j in range(len(order)):
    add_data_bars(ws, get_column_letter(j+2), mo_data_start, mo_data_end)

# Column widths
ws.column_dimensions['A'].width = 12
for j in range(len(order)):
    ws.column_dimensions[get_column_letter(j+2)].width = 16

ws.freeze_panes = f'B{yr_data_start}'

# Save
out = os.path.join(BASE, 'YEARLY_RETURNS_7STRATEGIES.xlsx')
wb.save(out)

# Patch XML for negative bar color (red)
tmpdir = tempfile.mkdtemp()
tmp_xlsx = os.path.join(tmpdir, 'patched.xlsx')
with zipfile.ZipFile(out, 'r') as zin:
    with zipfile.ZipFile(tmp_xlsx, 'w') as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                content = data.decode('utf-8')
                content = content.replace(
                    '</dataBar>',
                    '<negativeFillColor rgb="FFCC0000"/></dataBar>'
                )
                data = content.encode('utf-8')
            zout.writestr(item, data)
shutil.move(tmp_xlsx, out)
shutil.rmtree(tmpdir)

print(f"Saved: {out}")
print(f"Yearly: rows {yr_data_start}-{yr_data_end}")
print(f"Monthly: rows {mo_data_start}-{mo_data_end}")
