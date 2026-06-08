"""
5 Strategies Yearly Returns - Formatted Excel
5戦略の年次リターン比較
"""
import pandas as pd
import numpy as np
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import DataBar, FormatObject, Rule, CellIsRule
from openpyxl.styles import Color
from openpyxl.worksheet.page import PageMargins

from backtest_engine import (
    load_data, strategy_dd_vt_volspike, run_backtest
)
from test_ens2_strategies import strategy_ens2_asym_slope


def setup_print_settings(ws, orientation='portrait'):
    """印刷設定"""
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True


def main():
    print("=" * 100)
    print("5 Strategies Yearly Returns Generation")
    print("=" * 100)

    # Load data
    data_path = r"C:\Users\user\Desktop\nasdaq_backtest\NASDAQ_Dairy_since1973.csv"
    df = load_data(data_path)
    print(f"Data: {df['Date'].min().date()} to {df['Date'].max().date()}\n")

    close = df['Close']
    returns = close.pct_change()
    dates = df['Date']

    # ==========================================================================
    # Run all 5 strategies
    # ==========================================================================
    print("Running strategies...")

    # 1. Ens2(Asym+Slope) max_lev=3.0
    lev_ens2_3, pos_ens2_3 = strategy_ens2_asym_slope(close, returns, 0.82, 0.92, 0.25, 20, 5, 3.0)
    nav_ens2_3, _ = run_backtest(close, lev_ens2_3)
    print("  1. Ens2(Asym+Slope) max_lev=3.0 - Done")

    # 2. Ens2(Asym+Slope) max_lev=1.0
    lev_ens2_1, pos_ens2_1 = strategy_ens2_asym_slope(close, returns, 0.82, 0.92, 0.25, 20, 5, 1.0)
    nav_ens2_1, _ = run_backtest(close, lev_ens2_1)
    print("  2. Ens2(Asym+Slope) max_lev=1.0 - Done")

    # 3. DD+VT+VolSpike(1.5x)
    lev_volspike, pos_volspike = strategy_dd_vt_volspike(close, returns, 0.82, 0.92, 0.25, 10, 1.5)
    nav_volspike, _ = run_backtest(close, lev_volspike)
    print("  3. DD+VT+VolSpike(1.5x) - Done")

    # 4. BH 3x
    lev_bh = pd.Series(1.0, index=close.index)
    nav_bh, _ = run_backtest(close, lev_bh)
    print("  4. BH 3x - Done")

    # 5. NASDAQ 1x
    nasdaq_returns = returns.fillna(0)
    nav_nasdaq = (1 + nasdaq_returns).cumprod()
    print("  5. NASDAQ 1x - Done")

    # ==========================================================================
    # Calculate yearly returns
    # ==========================================================================
    print("\nCalculating yearly returns...")

    nav_df = pd.DataFrame({
        'Date': dates.values,
        'Ens2_3x': nav_ens2_3.values,
        'Ens2_1x': nav_ens2_1.values,
        'VolSpike': nav_volspike.values,
        'BH_3x': nav_bh.values,
        'NASDAQ_1x': nav_nasdaq.values
    })
    nav_df['Year'] = pd.to_datetime(nav_df['Date']).dt.year

    yearly_nav = nav_df.groupby('Year').last()

    # Calculate yearly returns
    yearly_data = []
    for col in ['Ens2_3x', 'Ens2_1x', 'VolSpike', 'BH_3x', 'NASDAQ_1x']:
        nav_series = yearly_nav[col]
        first_year_return = nav_series.iloc[0] - 1
        yoy_returns = nav_series.pct_change()
        yoy_returns.iloc[0] = first_year_return
        for year, ret in zip(yearly_nav.index, yoy_returns):
            existing = next((d for d in yearly_data if d['Year'] == year), None)
            if existing:
                existing[col] = ret * 100
            else:
                yearly_data.append({'Year': year, col: ret * 100})

    yearly_results = pd.DataFrame(yearly_data)

    # Year-end state for Ens2 (max_lev=1.0)
    pos_df = pd.DataFrame({'Date': dates.values, 'Position': pos_ens2_1.values})
    pos_df['Year'] = pd.to_datetime(pos_df['Date']).dt.year
    year_end_pos = pos_df.groupby('Year')['Position'].last()
    yearly_results['State'] = yearly_results['Year'].map(
        lambda y: '保有' if year_end_pos.get(y, 0) > 0.5 else '現金'
    )

    # ==========================================================================
    # Create Excel
    # ==========================================================================
    print("Creating Excel...")

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10, name='Yu Gothic')
    normal_font = Font(size=10, name='Yu Gothic')
    bold_font = Font(bold=True, size=10, name='Yu Gothic')
    title_font = Font(bold=True, size=14, name='Yu Gothic')
    section_font = Font(bold=True, size=11, color="4472C4", name='Yu Gothic')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    recommend_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb = Workbook()

    # ==========================================================================
    # Sheet 1: 年次リターン
    # ==========================================================================
    ws1 = wb.active
    ws1.title = "年次リターン"
    setup_print_settings(ws1, 'portrait')

    # Headers
    headers = [
        '年',
        'Ens2(Asym+Slope)\nmax_lev=3.0',
        'Ens2(Asym+Slope)\nmax_lev=1.0\n【推奨】',
        'DD+VT+VolSpike\n(1.5x)',
        'BH 3x',
        'NASDAQ 1x',
        '年末状態'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[1].height = 45

    # Data rows
    for row_idx, row in enumerate(yearly_results.itertuples(), 2):
        ws1.cell(row=row_idx, column=1, value=row.Year).border = thin_border
        ws1.cell(row=row_idx, column=2, value=round(row.Ens2_3x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=3, value=round(row.Ens2_1x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=4, value=round(row.VolSpike, 2)).border = thin_border
        ws1.cell(row=row_idx, column=5, value=round(row.BH_3x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=6, value=round(row.NASDAQ_1x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=7, value=row.State).border = thin_border

        for col in range(1, 8):
            ws1.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
            ws1.cell(row=row_idx, column=col).font = normal_font

    # Column widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 11
    ws1.column_dimensions['G'].width = 8

    # Highlight recommended column (C)
    for row in range(1, len(yearly_results) + 2):
        ws1.cell(row=row, column=3).fill = recommend_fill

    # Number format for return columns
    for col in range(2, 7):
        for row in range(2, len(yearly_results) + 2):
            ws1.cell(row=row, column=col).number_format = '+0.00;-0.00;0.00'

    last_row = len(yearly_results) + 1

    # Conditional formatting with data bars
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        # Cell background colors based on value
        green_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        green_font = Font(color='006100', name='Yu Gothic', size=10)
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font)
        )

        red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
        red_font = Font(color='9C0006', name='Yu Gothic', size=10)
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)
        )

        # Bidirectional data bar
        first = FormatObject(type='min')
        second = FormatObject(type='max')

        data_bar = DataBar(
            cfvo=[first, second],
            color=Color(rgb='5B9BD5'),
            showValue=True,
            minLength=0,
            maxLength=100
        )

        data_bar.negativeBarColor = Color(rgb='C00000')
        data_bar.negativeBarColorSameAsPositive = False
        data_bar.axisColor = Color(rgb='000000')
        data_bar.axisPosition = 'middle'
        data_bar.gradient = False

        rule = Rule(type='dataBar', dataBar=data_bar)
        ws1.conditional_formatting.add(f'{col_letter}2:{col_letter}{last_row}', rule)

    ws1.print_title_rows = '1:1'

    # ==========================================================================
    # Sheet 2: 戦略パラメータ
    # ==========================================================================
    ws2 = wb.create_sheet("戦略パラメータ")
    setup_print_settings(ws2, 'landscape')

    ws2.merge_cells('A1:E1')
    ws2.cell(row=1, column=1, value="戦略定義とパラメータ")
    ws2.cell(row=1, column=1).font = title_font

    params_headers = ['戦略名', 'max_lev', 'Layer1: DD制御', 'Layer2-4: 追加コンポーネント', '取引回数']
    for col, header in enumerate(params_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    strategies = [
        {
            'name': 'Ens2(Asym+Slope)',
            'max_lev': '3.0',
            'dd': 'Exit=-18%, Reentry=92%',
            'extra': 'AsymEWMA(up20/dn5) + VT(25%) + SlopeMult(MA200)',
            'trades': '30'
        },
        {
            'name': 'Ens2(Asym+Slope)\n【推奨】',
            'max_lev': '1.0',
            'dd': 'Exit=-18%, Reentry=92%',
            'extra': 'AsymEWMA(up20/dn5) + VT(25%) + SlopeMult(MA200)',
            'trades': '30'
        },
        {
            'name': 'DD+VT+VolSpike(1.5x)',
            'max_lev': '1.0',
            'dd': 'Exit=-18%, Reentry=92%',
            'extra': 'EWMA(10) + VT(25%) + VolSpike検出(1.5x)',
            'trades': '30'
        },
        {
            'name': 'Buy & Hold 3x',
            'max_lev': '-',
            'dd': 'なし（常時投資）',
            'extra': 'なし',
            'trades': '0'
        },
        {
            'name': 'NASDAQ 1x',
            'max_lev': '-',
            'dd': '該当なし（参考指数）',
            'extra': '該当なし',
            'trades': '0'
        }
    ]

    for row_idx, strat in enumerate(strategies, 4):
        ws2.cell(row=row_idx, column=1, value=strat['name']).border = thin_border
        ws2.cell(row=row_idx, column=2, value=strat['max_lev']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=strat['dd']).border = thin_border
        ws2.cell(row=row_idx, column=4, value=strat['extra']).border = thin_border
        ws2.cell(row=row_idx, column=5, value=strat['trades']).border = thin_border

        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws2.cell(row=row_idx, column=col).font = normal_font

        # Highlight recommended
        if '推奨' in strat['name']:
            for col in range(1, 6):
                ws2.cell(row=row_idx, column=col).fill = recommend_fill
                ws2.cell(row=row_idx, column=col).font = bold_font

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 10

    for row in range(4, 9):
        ws2.row_dimensions[row].height = 30

    # ==========================================================================
    # Sheet 3: サマリー統計
    # ==========================================================================
    ws3 = wb.create_sheet("サマリー統計")
    setup_print_settings(ws3, 'portrait')

    ws3.merge_cells('A1:F1')
    ws3.cell(row=1, column=1, value="パフォーマンスサマリー（1974-2021年）")
    ws3.cell(row=1, column=1).font = title_font

    # Calculate summary stats
    summary_headers = ['指標', 'Ens2\n(3.0)', 'Ens2\n(1.0)★', 'VolSpike', 'BH 3x', 'NASDAQ']
    for col, header in enumerate(summary_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws3.row_dimensions[3].height = 35

    # Metrics calculation
    def calc_stats(nav_series, yearly_col):
        years = len(nav_series) / 252
        final_nav = nav_series.iloc[-1]
        cagr = (final_nav ** (1/years)) - 1
        yearly_rets = yearly_results[yearly_col] / 100
        return {
            'final_nav': final_nav,
            'cagr': cagr,
            'best': yearly_rets.max(),
            'worst': yearly_rets.min(),
            'median': yearly_rets.median(),
            'positive': (yearly_rets > 0).sum(),
            'total': len(yearly_rets)
        }

    stats = {
        'Ens2_3x': calc_stats(nav_ens2_3, 'Ens2_3x'),
        'Ens2_1x': calc_stats(nav_ens2_1, 'Ens2_1x'),
        'VolSpike': calc_stats(nav_volspike, 'VolSpike'),
        'BH_3x': calc_stats(nav_bh, 'BH_3x'),
        'NASDAQ_1x': calc_stats(nav_nasdaq, 'NASDAQ_1x')
    }

    metrics_data = [
        ('最終NAV ($1開始)',
         f"${stats['Ens2_3x']['final_nav']:,.0f}",
         f"${stats['Ens2_1x']['final_nav']:,.0f}",
         f"${stats['VolSpike']['final_nav']:,.0f}",
         f"${stats['BH_3x']['final_nav']:,.0f}",
         f"${stats['NASDAQ_1x']['final_nav']:,.0f}"),
        ('CAGR',
         f"{stats['Ens2_3x']['cagr']*100:.2f}%",
         f"{stats['Ens2_1x']['cagr']*100:.2f}%",
         f"{stats['VolSpike']['cagr']*100:.2f}%",
         f"{stats['BH_3x']['cagr']*100:.2f}%",
         f"{stats['NASDAQ_1x']['cagr']*100:.2f}%"),
        ('最高年リターン',
         f"+{stats['Ens2_3x']['best']*100:.1f}%",
         f"+{stats['Ens2_1x']['best']*100:.1f}%",
         f"+{stats['VolSpike']['best']*100:.1f}%",
         f"+{stats['BH_3x']['best']*100:.1f}%",
         f"+{stats['NASDAQ_1x']['best']*100:.1f}%"),
        ('最低年リターン',
         f"{stats['Ens2_3x']['worst']*100:.1f}%",
         f"{stats['Ens2_1x']['worst']*100:.1f}%",
         f"{stats['VolSpike']['worst']*100:.1f}%",
         f"{stats['BH_3x']['worst']*100:.1f}%",
         f"{stats['NASDAQ_1x']['worst']*100:.1f}%"),
        ('中央値リターン',
         f"{stats['Ens2_3x']['median']*100:.1f}%",
         f"{stats['Ens2_1x']['median']*100:.1f}%",
         f"{stats['VolSpike']['median']*100:.1f}%",
         f"{stats['BH_3x']['median']*100:.1f}%",
         f"{stats['NASDAQ_1x']['median']*100:.1f}%"),
        ('プラス年数',
         f"{stats['Ens2_3x']['positive']}/{stats['Ens2_3x']['total']}",
         f"{stats['Ens2_1x']['positive']}/{stats['Ens2_1x']['total']}",
         f"{stats['VolSpike']['positive']}/{stats['VolSpike']['total']}",
         f"{stats['BH_3x']['positive']}/{stats['BH_3x']['total']}",
         f"{stats['NASDAQ_1x']['positive']}/{stats['NASDAQ_1x']['total']}"),
    ]

    for row_idx, row_data in enumerate(metrics_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')
            cell.font = bold_font if col_idx == 1 else normal_font

    # Highlight recommended column
    for row in range(3, 10):
        ws3.cell(row=row, column=3).fill = recommend_fill

    ws3.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 14

    # ==========================================================================
    # Sheet 4: 危機年パフォーマンス
    # ==========================================================================
    ws4 = wb.create_sheet("危機年パフォーマンス")
    setup_print_settings(ws4, 'portrait')

    ws4.merge_cells('A1:F1')
    ws4.cell(row=1, column=1, value="危機年パフォーマンス比較")
    ws4.cell(row=1, column=1).font = title_font

    crisis_headers = ['年', 'Ens2(3.0)', 'Ens2(1.0)★', 'VolSpike', 'BH 3x', 'NASDAQ']
    for col, header in enumerate(crisis_headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    crisis_years = [1974, 1987, 2000, 2001, 2002, 2008, 2011, 2020]
    row_idx = 4
    for year in crisis_years:
        year_data = yearly_results[yearly_results['Year'] == year]
        if len(year_data) > 0:
            row = year_data.iloc[0]
            ws4.cell(row=row_idx, column=1, value=year).border = thin_border
            ws4.cell(row=row_idx, column=2, value=f"{row['Ens2_3x']:.1f}%").border = thin_border
            ws4.cell(row=row_idx, column=3, value=f"{row['Ens2_1x']:.1f}%").border = thin_border
            ws4.cell(row=row_idx, column=4, value=f"{row['VolSpike']:.1f}%").border = thin_border
            ws4.cell(row=row_idx, column=5, value=f"{row['BH_3x']:.1f}%").border = thin_border
            ws4.cell(row=row_idx, column=6, value=f"{row['NASDAQ_1x']:.1f}%").border = thin_border

            for col in range(1, 7):
                ws4.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                ws4.cell(row=row_idx, column=col).font = normal_font

            # Highlight recommended column
            ws4.cell(row=row_idx, column=3).fill = recommend_fill

            row_idx += 1

    ws4.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws4.column_dimensions[col].width = 14

    # ==========================================================================
    # Save
    # ==========================================================================
    excel_path = r"C:\Users\user\Desktop\nasdaq_backtest\5Strategies_年次リターン.xlsx"
    wb.save(excel_path)

    print(f"\nExcel saved: {excel_path}")
    print("\nSheets:")
    print("  1. 年次リターン - 5戦略の年次リターン比較")
    print("  2. 戦略パラメータ - 各戦略の定義")
    print("  3. サマリー統計 - パフォーマンス比較")
    print("  4. 危機年パフォーマンス - 危機年の比較")

    # Print summary table
    print("\n" + "=" * 100)
    print("YEARLY RETURNS SUMMARY")
    print("=" * 100)
    print(f"\n{'Year':<6} {'Ens2(3.0)':>12} {'Ens2(1.0)':>12} {'VolSpike':>12} {'BH 3x':>12} {'NASDAQ':>10}")
    print("-" * 70)
    for _, row in yearly_results.iterrows():
        print(f"{int(row['Year']):<6} {row['Ens2_3x']:>+11.2f}% {row['Ens2_1x']:>+11.2f}% "
              f"{row['VolSpike']:>+11.2f}% {row['BH_3x']:>+11.2f}% {row['NASDAQ_1x']:>+9.2f}%")

    return yearly_results


if __name__ == "__main__":
    results = main()
