"""
DD+VT+VolSpike(1.5x) - Japanese Excel Report (Print-Friendly)
"""
import pandas as pd
import numpy as np
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import DataBarRule, CellIsRule, ColorScaleRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from backtest_engine import (
    load_data, strategy_dd_vt_volspike, strategy_baseline_dd_vt,
    run_backtest
)

def setup_print_settings(ws, orientation='portrait'):
    """印刷設定"""
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True

def main():
    print("=" * 80)
    print("日本語Excel生成中...")
    print("=" * 80)

    # Load data
    data_path = r"C:\Users\user\Desktop\nasdaq_backtest\NASDAQ_Dairy_since1973.csv"
    df = load_data(data_path)
    close = df['Close']
    returns = close.pct_change()
    dates = df['Date']

    # Run strategies
    lev_volspike, pos_volspike = strategy_dd_vt_volspike(close, returns, 0.82, 0.92, 0.25, 10, 1.5)
    nav_volspike, _ = run_backtest(close, lev_volspike)

    lev_baseline, pos_baseline = strategy_baseline_dd_vt(close, returns, 0.82, 0.92, 0.25, 10)
    nav_baseline, _ = run_backtest(close, lev_baseline)

    lev_bh = pd.Series(1.0, index=close.index)
    nav_bh, _ = run_backtest(close, lev_bh)

    nasdaq_returns = returns.fillna(0)
    nav_nasdaq = (1 + nasdaq_returns).cumprod()

    # Calculate yearly returns
    nav_df = pd.DataFrame({
        'Date': dates.values,
        'VolSpike': nav_volspike.values,
        'Baseline': nav_baseline.values,
        'BH_3x': nav_bh.values,
        'NASDAQ_1x': nav_nasdaq.values
    })
    nav_df['Year'] = pd.to_datetime(nav_df['Date']).dt.year
    yearly_nav = nav_df.groupby('Year').last()

    yearly_data = []
    for col in ['VolSpike', 'Baseline', 'BH_3x', 'NASDAQ_1x']:
        nav_series = yearly_nav[col]
        first_year_return = nav_series.iloc[0] - 1
        yoy_returns = nav_series.pct_change()
        yoy_returns.iloc[0] = first_year_return
        for year, ret in zip(yearly_nav.index, yoy_returns):
            existing = next((d for d in yearly_data if d['Year'] == year), None)
            if existing:
                existing[col] = ret * 100
            else:
                yearly_data.append({'Year': year, col: ret * 100})

    yearly_results = pd.DataFrame(yearly_data)

    pos_df = pd.DataFrame({'Date': dates.values, 'Position': pos_volspike.values})
    pos_df['Year'] = pd.to_datetime(pos_df['Date']).dt.year
    year_end_pos = pos_df.groupby('Year')['Position'].last()
    yearly_results['State'] = yearly_results['Year'].map(
        lambda y: '保有' if year_end_pos.get(y, 0) > 0.5 else '現金'
    )

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10, name='Yu Gothic')
    normal_font = Font(size=10, name='Yu Gothic')
    bold_font = Font(bold=True, size=10, name='Yu Gothic')
    title_font = Font(bold=True, size=14, name='Yu Gothic')
    section_font = Font(bold=True, size=11, color="4472C4", name='Yu Gothic')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()

    # ==========================================================================
    # Sheet 1: 年次リターン
    # ==========================================================================
    ws1 = wb.active
    ws1.title = "年次リターン"
    setup_print_settings(ws1, 'portrait')

    headers = ['年', 'DD+VT+VolSpike\n(1.5x) [推奨]', 'DD(-18/92)\n+VT(25%)', 'Buy&Hold\n3倍', 'NASDAQ\n1倍', '年末\n状態']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[1].height = 35

    for row_idx, row in enumerate(yearly_results.itertuples(), 2):
        ws1.cell(row=row_idx, column=1, value=row.Year).border = thin_border
        ws1.cell(row=row_idx, column=2, value=round(row.VolSpike, 2)).border = thin_border
        ws1.cell(row=row_idx, column=3, value=round(row.Baseline, 2)).border = thin_border
        ws1.cell(row=row_idx, column=4, value=round(row.BH_3x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=5, value=round(row.NASDAQ_1x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=6, value=row.State).border = thin_border

        for col in range(1, 7):
            ws1.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
            ws1.cell(row=row_idx, column=col).font = normal_font

    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 11
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 7

    for col in range(2, 6):
        for row in range(2, len(yearly_results) + 2):
            ws1.cell(row=row, column=col).number_format = '+0.00;-0.00;0.00'

    last_row = len(yearly_results) + 1

    from openpyxl.formatting.rule import DataBar, FormatObject, Rule
    from openpyxl.styles import Color

    for col_letter in ['B', 'C', 'D', 'E']:
        # Cell background colors based on value
        green_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        green_font = Font(color='006100', name='Yu Gothic', size=10)
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font)
        )

        red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
        red_font = Font(color='9C0006', name='Yu Gothic', size=10)
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)
        )

        # Bidirectional data bar: negative=left(red), positive=right(blue), solid color
        # Create format objects for min/max
        first = FormatObject(type='min')
        second = FormatObject(type='max')

        # Create data bar with bidirectional axis (middle)
        data_bar = DataBar(
            cfvo=[first, second],
            color=Color(rgb='5B9BD5'),  # Blue for positive
            showValue=True,
            minLength=0,
            maxLength=100
        )

        # Set negative bar color (red) and axis to middle for bidirectional
        data_bar.negativeBarColor = Color(rgb='C00000')
        data_bar.negativeBarColorSameAsPositive = False
        data_bar.axisColor = Color(rgb='000000')
        data_bar.axisPosition = 'middle'

        # Disable gradient (solid fill)
        data_bar.gradient = False

        # Create rule and add
        rule = Rule(type='dataBar', dataBar=data_bar)
        ws1.conditional_formatting.add(f'{col_letter}2:{col_letter}{last_row}', rule)

    ws1.print_title_rows = '1:1'

    # ==========================================================================
    # Sheet 2: 戦略パラメータ
    # ==========================================================================
    ws2 = wb.create_sheet("戦略パラメータ")
    setup_print_settings(ws2, 'landscape')

    ws2.merge_cells('A1:E1')
    ws2.cell(row=1, column=1, value="戦略定義とパラメータ")
    ws2.cell(row=1, column=1).font = title_font

    params_headers = ['戦略名', 'Layer1: DD制御', 'Layer2: ボラティリティ\nターゲティング', 'Layer3: 追加フィルター', '取引\n回数']
    for col, header in enumerate(params_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws2.row_dimensions[3].height = 35

    strategies = [
        {
            'name': 'DD(-18/92)+VT(25%)\n+VolSpike(1.5x)\n【推奨戦略】',
            'dd': '退出: 200日高値から-18%下落\n復帰: 92%回復時',
            'vt': 'Target Vol: 25%\nEWMA Span: 10日\nmax_lev: 1.0',
            'filter': 'ボラ急騰検出:\n今日Vol/昨日Vol > 1.5倍\n→ レバレッジ×0.5',
            'trades': '30回'
        },
        {
            'name': 'DD(-18/92)+VT(25%)\n【ベースライン】',
            'dd': '退出: 200日高値から-18%下落\n復帰: 92%回復時',
            'vt': 'Target Vol: 25%\nEWMA Span: 10日\nmax_lev: 1.0',
            'filter': 'なし',
            'trades': '30回'
        },
        {
            'name': 'Buy & Hold 3倍\n【参考】',
            'dd': 'なし（常時投資）',
            'vt': 'なし（常時100%）',
            'filter': 'なし',
            'trades': '0回'
        },
        {
            'name': 'NASDAQ 1倍\n【参考指数】',
            'dd': '該当なし',
            'vt': '該当なし',
            'filter': '該当なし（レバなし指数）',
            'trades': '0回'
        }
    ]

    for row_idx, strat in enumerate(strategies, 4):
        ws2.cell(row=row_idx, column=1, value=strat['name']).border = thin_border
        ws2.cell(row=row_idx, column=2, value=strat['dd']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=strat['vt']).border = thin_border
        ws2.cell(row=row_idx, column=4, value=strat['filter']).border = thin_border
        ws2.cell(row=row_idx, column=5, value=strat['trades']).border = thin_border
        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws2.cell(row=row_idx, column=col).font = normal_font
        ws2.cell(row=row_idx, column=1).font = bold_font

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 26
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 26
    ws2.column_dimensions['E'].width = 8

    for row in range(4, 8):
        ws2.row_dimensions[row].height = 55

    # ==========================================================================
    # Sheet 3: 計算前提・戦略ロジック（統合）
    # ==========================================================================
    ws3 = wb.create_sheet("計算前提・ロジック詳細")
    setup_print_settings(ws3, 'portrait')

    row = 1

    # Title
    ws3.merge_cells(f'A{row}:B{row}')
    ws3.cell(row=row, column=1, value="計算前提・戦略ロジック詳細")
    ws3.cell(row=row, column=1).font = title_font
    row += 2

    # Section 1: 計算前提
    ws3.merge_cells(f'A{row}:B{row}')
    ws3.cell(row=row, column=1, value="■ 計算前提")
    ws3.cell(row=row, column=1).font = section_font
    row += 1

    assumptions = [
        ("データ期間", "1974年1月2日 ～ 2021年5月7日（47年間、11,943営業日）"),
        ("データソース", "NASDAQ Composite Index (^IXIC) - Yahoo Finance"),
        ("ベースレバレッジ", "3倍（日次リバランス、TQQQシミュレーション）"),
        ("年間コスト", "0.9%（経費率、投資時のみ日割りで控除）"),
        ("現金保有時リターン", "0%（金利なし）"),
        ("無リスク金利", "0%（シャープレシオ計算用）"),
        ("取引回数カウント", "バイナリ状態遷移のみ（保有↔現金）"),
        ("VT max_lev", "1.0（実効レバレッジ上限3倍、9倍ではない）"),
        ("EWMAボラティリティ", "指数加重移動平均、Span=10日、年率換算（×√252）"),
        ("DDルックバック", "200営業日で高値追跡"),
        ("Vol Spike閾値", "今日のEWMA Vol / 昨日のEWMA Vol > 1.5倍で発動"),
    ]

    for param, value in assumptions:
        ws3.cell(row=row, column=1, value=param).font = bold_font
        ws3.cell(row=row, column=1).border = thin_border
        ws3.cell(row=row, column=2, value=value).font = normal_font
        ws3.cell(row=row, column=2).border = thin_border
        row += 1

    row += 1

    # Section 2: 戦略ロジック
    ws3.merge_cells(f'A{row}:B{row}')
    ws3.cell(row=row, column=1, value="■ DD+VT+VolSpike(1.5x) 戦略ロジック")
    ws3.cell(row=row, column=1).font = section_font
    row += 1

    logic_content = [
        ("【Layer 1: ドローダウン(DD)制御】", ""),
        ("目的", "暴落時の壊滅的損失を回避"),
        ("仕組み", "過去200日の最高値を追跡"),
        ("退出シグナル", "現在価格 / 200日高値 ≤ 0.82 → 現金へ退避"),
        ("復帰シグナル", "現在価格 / 200日高値 ≥ 0.92 → 保有へ復帰"),
        ("", ""),
        ("【Layer 2: ボラティリティターゲティング(VT)】", ""),
        ("目的", "市場ボラティリティに応じてポジションサイズ調整"),
        ("EWMA Vol計算", "ewma_vol = EWMA(日次リターン, span=10) × √252"),
        ("レバレッジ計算", "vt_leverage = min(0.25 / ewma_vol, 1.0)"),
        ("効果", "高ボラ→低レバ、低ボラ→高レバ（上限1.0）"),
        ("", ""),
        ("【Layer 3: ボラ急騰検出】", ""),
        ("目的", "急激なボラティリティ上昇時の追加保護"),
        ("検出方法", "vol_ratio = 今日のewma_vol / 昨日のewma_vol"),
        ("発動条件", "vol_ratio > 1.5 → スパイク検出"),
        ("アクション", "スパイク時: final_leverage = vt_leverage × 0.5"),
        ("", ""),
        ("【最終ポジション計算】", ""),
        ("計算式", "final_lev = DD信号 × VTレバ × (0.5 if スパイク else 1.0)"),
        ("レンジ", "0（現金）～ 1.0（3倍商品に100%投資）"),
        ("実効レバレッジ", "0倍 ～ 3倍（final_leverage × 3）"),
    ]

    for label, desc in logic_content:
        if label.startswith("【"):
            ws3.merge_cells(f'A{row}:B{row}')
            ws3.cell(row=row, column=1, value=label).font = Font(bold=True, size=11, color="2F5496", name='Yu Gothic')
        elif label == "":
            pass
        else:
            ws3.cell(row=row, column=1, value=label).font = bold_font
            ws3.cell(row=row, column=1).border = thin_border
            ws3.cell(row=row, column=2, value=desc).font = normal_font
            ws3.cell(row=row, column=2).border = thin_border
        row += 1

    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 55

    # ==========================================================================
    # Sheet 4: サマリー統計
    # ==========================================================================
    ws4 = wb.create_sheet("サマリー統計")
    setup_print_settings(ws4, 'portrait')

    ws4.merge_cells('A1:E1')
    ws4.cell(row=1, column=1, value="パフォーマンスサマリー（1974-2021年）")
    ws4.cell(row=1, column=1).font = title_font

    summary_headers = ['指標', 'DD+VT+VolSpike\n(1.5x) [推奨]', 'DD(-18/92)\n+VT(25%)', 'Buy&Hold\n3倍', 'NASDAQ\n1倍']
    for col, header in enumerate(summary_headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws4.row_dimensions[3].height = 35

    metrics_data = [
        ('最終資産（$1開始）',
         f"${nav_volspike.iloc[-1]:,.0f}",
         f"${nav_baseline.iloc[-1]:,.0f}",
         f"${nav_bh.iloc[-1]:,.0f}",
         f"${nav_nasdaq.iloc[-1]:,.0f}"),
        ('年率複利成長率(CAGR)',
         f"{(nav_volspike.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_baseline.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_bh.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_nasdaq.iloc[-1]**(1/47)-1)*100:.2f}%"),
        ('最高年リターン',
         f"+{yearly_results['VolSpike'].max():.1f}%",
         f"+{yearly_results['Baseline'].max():.1f}%",
         f"+{yearly_results['BH_3x'].max():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].max():.1f}%"),
        ('最低年リターン',
         f"{yearly_results['VolSpike'].min():.1f}%",
         f"{yearly_results['Baseline'].min():.1f}%",
         f"{yearly_results['BH_3x'].min():.1f}%",
         f"{yearly_results['NASDAQ_1x'].min():.1f}%"),
        ('プラス年数',
         f"{(yearly_results['VolSpike'] > 0).sum()}/48年",
         f"{(yearly_results['Baseline'] > 0).sum()}/48年",
         f"{(yearly_results['BH_3x'] > 0).sum()}/48年",
         f"{(yearly_results['NASDAQ_1x'] > 0).sum()}/48年"),
        ('勝率',
         f"{(yearly_results['VolSpike'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['Baseline'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['BH_3x'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['NASDAQ_1x'] > 0).mean()*100:.1f}%"),
        ('平均年リターン',
         f"+{yearly_results['VolSpike'].mean():.1f}%",
         f"+{yearly_results['Baseline'].mean():.1f}%",
         f"+{yearly_results['BH_3x'].mean():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].mean():.1f}%"),
        ('中央値年リターン',
         f"+{yearly_results['VolSpike'].median():.1f}%",
         f"+{yearly_results['Baseline'].median():.1f}%",
         f"+{yearly_results['BH_3x'].median():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].median():.1f}%"),
    ]

    for row_idx, row_data in enumerate(metrics_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')
            cell.font = bold_font if col_idx == 1 else normal_font

    # Highlight recommended strategy column
    highlight_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    for row in range(3, 12):
        ws4.cell(row=row, column=2).fill = highlight_fill

    ws4.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws4.column_dimensions[col].width = 18

    # Crisis years section
    row = 14
    ws4.merge_cells(f'A{row}:E{row}')
    ws4.cell(row=row, column=1, value="危機年パフォーマンス比較")
    ws4.cell(row=row, column=1).font = section_font
    row += 1

    crisis_headers = ['年', 'DD+VT+VolSpike', 'Buy&Hold 3x', 'NASDAQ 1x', '保護効果']
    for col, header in enumerate(crisis_headers, 1):
        cell = ws4.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    crisis_years = [1974, 1987, 2000, 2001, 2002, 2008]
    for year in crisis_years:
        year_data = yearly_results[yearly_results['Year'] == year].iloc[0]
        protection = year_data['VolSpike'] - year_data['BH_3x']

        ws4.cell(row=row, column=1, value=year).border = thin_border
        ws4.cell(row=row, column=2, value=f"{year_data['VolSpike']:.1f}%").border = thin_border
        ws4.cell(row=row, column=3, value=f"{year_data['BH_3x']:.1f}%").border = thin_border
        ws4.cell(row=row, column=4, value=f"{year_data['NASDAQ_1x']:.1f}%").border = thin_border
        ws4.cell(row=row, column=5, value=f"+{protection:.1f}%").border = thin_border

        for col in range(1, 6):
            ws4.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            ws4.cell(row=row, column=col).font = normal_font

        # Highlight protection
        ws4.cell(row=row, column=5).fill = highlight_fill
        ws4.cell(row=row, column=5).font = Font(bold=True, color='006100', name='Yu Gothic', size=10)

        row += 1

    # ==========================================================================
    # Save
    # ==========================================================================
    excel_path = r"C:\Users\user\Desktop\nasdaq_backtest\VolSpike_年次リターン_v2.xlsx"
    wb.save(excel_path)

    print(f"\n保存完了: {excel_path}")
    print("\nシート構成:")
    print("  1. 年次リターン - 条件付き書式付き年次リターン表")
    print("  2. 戦略パラメータ - 各戦略の定義とパラメータ")
    print("  3. 計算前提・ロジック詳細 - 計算前提と戦略ロジック（統合）")
    print("  4. サマリー統計 - パフォーマンス比較と危機年分析")

if __name__ == "__main__":
    main()
