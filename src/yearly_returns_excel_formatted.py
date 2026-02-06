"""
DD+VT+VolSpike(1.5x) - Formatted Excel with Conditional Formatting
"""
import pandas as pd
import numpy as np
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from backtest_engine import (
    load_data, strategy_dd_vt_volspike, strategy_baseline_dd_vt,
    run_backtest
)

def main():
    print("=" * 80)
    print("Generating Formatted Excel Report")
    print("=" * 80)

    # Load data
    data_path = r"C:\Users\user\Desktop\nasdaq_backtest\NASDAQ_Dairy_since1973.csv"
    df = load_data(data_path)
    close = df['Close']
    returns = close.pct_change()
    dates = df['Date']

    # ==========================================================================
    # Run strategies
    # ==========================================================================

    # Strategy 1: DD(-18/92)+VT(25%)+VolSpike(1.5x) [Top Recommended]
    lev_volspike, pos_volspike = strategy_dd_vt_volspike(
        close, returns, 0.82, 0.92, 0.25, 10, 1.5
    )
    nav_volspike, _ = run_backtest(close, lev_volspike)

    # Strategy 2: DD(-18/92)+VT(25%) Baseline
    lev_baseline, pos_baseline = strategy_baseline_dd_vt(
        close, returns, 0.82, 0.92, 0.25, 10
    )
    nav_baseline, _ = run_backtest(close, lev_baseline)

    # Strategy 3: Buy & Hold 3x
    lev_bh = pd.Series(1.0, index=close.index)
    nav_bh, _ = run_backtest(close, lev_bh)

    # Strategy 4: NASDAQ 1x
    nasdaq_returns = returns.fillna(0)
    nav_nasdaq = (1 + nasdaq_returns).cumprod()

    # ==========================================================================
    # Calculate yearly returns
    # ==========================================================================

    nav_df = pd.DataFrame({
        'Date': dates.values,
        'VolSpike': nav_volspike.values,
        'Baseline': nav_baseline.values,
        'BH_3x': nav_bh.values,
        'NASDAQ_1x': nav_nasdaq.values
    })
    nav_df['Year'] = pd.to_datetime(nav_df['Date']).dt.year

    yearly_nav = nav_df.groupby('Year').last()

    # Calculate yearly returns
    yearly_data = []

    for col in ['VolSpike', 'Baseline', 'BH_3x', 'NASDAQ_1x']:
        nav_series = yearly_nav[col]
        first_year_return = nav_series.iloc[0] - 1
        yoy_returns = nav_series.pct_change()
        yoy_returns.iloc[0] = first_year_return

        for year, ret in zip(yearly_nav.index, yoy_returns):
            existing = next((d for d in yearly_data if d['Year'] == year), None)
            if existing:
                existing[col] = ret * 100
            else:
                yearly_data.append({'Year': year, col: ret * 100})

    yearly_results = pd.DataFrame(yearly_data)

    # Add year-end state
    pos_df = pd.DataFrame({
        'Date': dates.values,
        'Position': pos_volspike.values
    })
    pos_df['Year'] = pd.to_datetime(pos_df['Date']).dt.year
    year_end_pos = pos_df.groupby('Year')['Position'].last()
    yearly_results['State'] = yearly_results['Year'].map(
        lambda y: 'HOLD' if year_end_pos.get(y, 0) > 0.5 else 'CASH'
    )

    # ==========================================================================
    # Create Excel Workbook
    # ==========================================================================

    wb = Workbook()

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==========================================================================
    # Sheet 1: Yearly Returns
    # ==========================================================================

    ws1 = wb.active
    ws1.title = "Yearly_Returns"

    # Headers
    headers = ['Year', 'DD+VT+VolSpike(1.5x)', 'DD(-18/92)+VT(25%)', 'Buy&Hold 3x', 'NASDAQ 1x', 'Year-End State']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data
    for row_idx, row in enumerate(yearly_results.itertuples(), 2):
        ws1.cell(row=row_idx, column=1, value=row.Year).border = thin_border
        ws1.cell(row=row_idx, column=2, value=round(row.VolSpike, 2)).border = thin_border
        ws1.cell(row=row_idx, column=3, value=round(row.Baseline, 2)).border = thin_border
        ws1.cell(row=row_idx, column=4, value=round(row.BH_3x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=5, value=round(row.NASDAQ_1x, 2)).border = thin_border
        ws1.cell(row=row_idx, column=6, value=row.State).border = thin_border

        # Center alignment
        for col in range(1, 7):
            ws1.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

    # Column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14

    # Number format for return columns
    for col in range(2, 6):
        for row in range(2, len(yearly_results) + 2):
            ws1.cell(row=row, column=col).number_format = '+0.00;-0.00;0.00'

    # Conditional formatting - Data Bars for each return column
    last_row = len(yearly_results) + 1

    # Green data bar for positive, Red for negative
    for col_letter in ['B', 'C', 'D', 'E']:
        # Data bar rule (blue gradient)
        data_bar_rule = DataBarRule(
            start_type='num', start_value=-100,
            end_type='num', end_value=400,
            color='638EC6',
            showValue=True,
            minLength=None,
            maxLength=None
        )
        ws1.conditional_formatting.add(f'{col_letter}2:{col_letter}{last_row}', data_bar_rule)

        # Green fill for positive
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        green_font = Font(color='006100')
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font)
        )

        # Red fill for negative
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006')
        ws1.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)
        )

    # ==========================================================================
    # Sheet 2: Strategy Parameters
    # ==========================================================================

    ws2 = wb.create_sheet("Strategy_Parameters")

    # Title
    ws2.merge_cells('A1:E1')
    ws2.cell(row=1, column=1, value="Strategy Definitions & Calculation Assumptions")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Strategy parameters table
    params_headers = ['Strategy Name', 'Layer 1: DD Control', 'Layer 2: Volatility Targeting', 'Layer 3: Additional Filter', 'Trade Count (47yr)']
    for col, header in enumerate(params_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    strategies = [
        {
            'name': 'DD(-18/92)+VT(25%)+VolSpike(1.5x)',
            'dd': 'Exit: -18% from 200d peak\nReentry: 92% recovery',
            'vt': 'Target Vol: 25%\nEWMA Span: 10\nmax_lev: 1.0',
            'filter': 'Vol Spike Detection:\nIf today_vol/yesterday_vol > 1.5x\n→ Leverage × 0.5',
            'trades': '30'
        },
        {
            'name': 'DD(-18/92)+VT(25%)\n[Baseline]',
            'dd': 'Exit: -18% from 200d peak\nReentry: 92% recovery',
            'vt': 'Target Vol: 25%\nEWMA Span: 10\nmax_lev: 1.0',
            'filter': 'None',
            'trades': '30'
        },
        {
            'name': 'Buy & Hold 3x',
            'dd': 'None (always invested)',
            'vt': 'None (always 100%)',
            'filter': 'None',
            'trades': '0'
        },
        {
            'name': 'NASDAQ 1x\n[Reference]',
            'dd': 'N/A',
            'vt': 'N/A',
            'filter': 'N/A (unleveraged index)',
            'trades': '0'
        }
    ]

    for row_idx, strat in enumerate(strategies, 4):
        ws2.cell(row=row_idx, column=1, value=strat['name']).border = thin_border
        ws2.cell(row=row_idx, column=2, value=strat['dd']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=strat['vt']).border = thin_border
        ws2.cell(row=row_idx, column=4, value=strat['filter']).border = thin_border
        ws2.cell(row=row_idx, column=5, value=strat['trades']).border = thin_border

        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 16

    for row in range(4, 8):
        ws2.row_dimensions[row].height = 60

    # ==========================================================================
    # Sheet 3: Calculation Assumptions
    # ==========================================================================

    ws3 = wb.create_sheet("Calculation_Assumptions")

    assumptions = [
        ("Data Period", "1974-01-02 to 2021-05-07 (47 years, 11,943 trading days)"),
        ("Data Source", "NASDAQ Composite Index (^IXIC) - Yahoo Finance"),
        ("Base Leverage", "3x daily rebalancing (simulating TQQQ-like product)"),
        ("Annual Cost", "0.9% (expense ratio, charged daily when invested)"),
        ("Cash Return", "0% (no interest earned during CASH periods)"),
        ("Risk-Free Rate", "0% (for Sharpe ratio calculation)"),
        ("Trade Counting", "Binary state changes only (HOLD↔CASH transitions)"),
        ("VT max_lev", "1.0 (effective leverage capped at 3x, not 9x)"),
        ("EWMA Volatility", "Exponentially Weighted Moving Average, Span=10 days, annualized (×√252)"),
        ("DD Lookback", "200 trading days for peak tracking"),
        ("Vol Spike", "Compares today's EWMA vol to yesterday's; if ratio > 1.5x, halve leverage"),
    ]

    ws3.merge_cells('A1:B1')
    ws3.cell(row=1, column=1, value="Calculation Assumptions & Parameters")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    for col, header in enumerate(['Parameter', 'Value / Description'], 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, (param, value) in enumerate(assumptions, 4):
        ws3.cell(row=row_idx, column=1, value=param).border = thin_border
        ws3.cell(row=row_idx, column=2, value=value).border = thin_border
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 70

    # ==========================================================================
    # Sheet 4: DD+VT+VolSpike Logic (Detailed)
    # ==========================================================================

    ws4 = wb.create_sheet("Strategy_Logic_Detail")

    ws4.merge_cells('A1:B1')
    ws4.cell(row=1, column=1, value="DD+VT+VolSpike(1.5x) Strategy - Detailed Logic")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=14)

    logic_content = [
        ("", ""),
        ("【Layer 1: Drawdown (DD) Control】", ""),
        ("Purpose", "Avoid catastrophic losses during market crashes"),
        ("Mechanism", "Track rolling 200-day maximum price"),
        ("Exit Signal", "If current_price / peak_200d <= 0.82 → Switch to CASH"),
        ("Reentry Signal", "If current_price / peak_200d >= 0.92 → Switch to HOLD"),
        ("", ""),
        ("【Layer 2: Volatility Targeting (VT)】", ""),
        ("Purpose", "Adjust position size based on market volatility"),
        ("EWMA Vol", "ewma_vol = EWMA(daily_returns, span=10) × √252"),
        ("Leverage Calc", "vt_leverage = min(0.25 / ewma_vol, 1.0)"),
        ("Effect", "High vol → lower leverage, Low vol → higher leverage (capped at 1.0)"),
        ("", ""),
        ("【Layer 3: Vol Spike Detection】", ""),
        ("Purpose", "Additional protection against sudden volatility increases"),
        ("Detection", "vol_ratio = today_ewma_vol / yesterday_ewma_vol"),
        ("Condition", "If vol_ratio > 1.5 → Spike detected"),
        ("Action", "If spike: final_leverage = vt_leverage × 0.5"),
        ("", ""),
        ("【Final Position Calculation】", ""),
        ("Formula", "final_leverage = dd_signal × vt_leverage × (0.5 if vol_spike else 1.0)"),
        ("Range", "0 (CASH) to 1.0 (max position in 3x product)"),
        ("Effective Lev", "0x to 3x (final_leverage × 3)"),
    ]

    for row_idx, (label, desc) in enumerate(logic_content, 3):
        cell1 = ws4.cell(row=row_idx, column=1, value=label)
        cell2 = ws4.cell(row=row_idx, column=2, value=desc)

        if label.startswith("【"):
            cell1.font = Font(bold=True, size=12, color="4472C4")
            ws4.merge_cells(f'A{row_idx}:B{row_idx}')
        elif label:
            cell1.font = Font(bold=True)

    ws4.column_dimensions['A'].width = 18
    ws4.column_dimensions['B'].width = 65

    # ==========================================================================
    # Sheet 5: Summary Statistics
    # ==========================================================================

    ws5 = wb.create_sheet("Summary_Statistics")

    ws5.merge_cells('A1:E1')
    ws5.cell(row=1, column=1, value="Performance Summary (1974-2021)")
    ws5.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Calculate summary stats
    summary_headers = ['Metric', 'DD+VT+VolSpike(1.5x)', 'DD(-18/92)+VT(25%)', 'Buy&Hold 3x', 'NASDAQ 1x']
    for col, header in enumerate(summary_headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Metrics
    metrics_data = [
        ('Final NAV ($1 start)',
         f"${nav_volspike.iloc[-1]:,.0f}",
         f"${nav_baseline.iloc[-1]:,.0f}",
         f"${nav_bh.iloc[-1]:,.0f}",
         f"${nav_nasdaq.iloc[-1]:,.0f}"),
        ('CAGR',
         f"{(nav_volspike.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_baseline.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_bh.iloc[-1]**(1/47)-1)*100:.2f}%",
         f"{(nav_nasdaq.iloc[-1]**(1/47)-1)*100:.2f}%"),
        ('Best Year',
         f"+{yearly_results['VolSpike'].max():.1f}%",
         f"+{yearly_results['Baseline'].max():.1f}%",
         f"+{yearly_results['BH_3x'].max():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].max():.1f}%"),
        ('Worst Year',
         f"{yearly_results['VolSpike'].min():.1f}%",
         f"{yearly_results['Baseline'].min():.1f}%",
         f"{yearly_results['BH_3x'].min():.1f}%",
         f"{yearly_results['NASDAQ_1x'].min():.1f}%"),
        ('Positive Years',
         f"{(yearly_results['VolSpike'] > 0).sum()}/48",
         f"{(yearly_results['Baseline'] > 0).sum()}/48",
         f"{(yearly_results['BH_3x'] > 0).sum()}/48",
         f"{(yearly_results['NASDAQ_1x'] > 0).sum()}/48"),
        ('Win Rate',
         f"{(yearly_results['VolSpike'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['Baseline'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['BH_3x'] > 0).mean()*100:.1f}%",
         f"{(yearly_results['NASDAQ_1x'] > 0).mean()*100:.1f}%"),
        ('Avg Annual Return',
         f"+{yearly_results['VolSpike'].mean():.1f}%",
         f"+{yearly_results['Baseline'].mean():.1f}%",
         f"+{yearly_results['BH_3x'].mean():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].mean():.1f}%"),
        ('Median Annual Return',
         f"+{yearly_results['VolSpike'].median():.1f}%",
         f"+{yearly_results['Baseline'].median():.1f}%",
         f"+{yearly_results['BH_3x'].median():.1f}%",
         f"+{yearly_results['NASDAQ_1x'].median():.1f}%"),
    ]

    for row_idx, row_data in enumerate(metrics_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')
            if col_idx == 1:
                cell.font = Font(bold=True)

    ws5.column_dimensions['A'].width = 22
    for col in ['B', 'C', 'D', 'E']:
        ws5.column_dimensions[col].width = 22

    # ==========================================================================
    # Save Excel
    # ==========================================================================

    excel_path = r"C:\Users\user\Desktop\nasdaq_backtest\VolSpike_Yearly_Returns.xlsx"
    wb.save(excel_path)

    print(f"\nExcel saved to: {excel_path}")
    print("\nSheets created:")
    print("  1. Yearly_Returns - Annual returns with conditional formatting")
    print("  2. Strategy_Parameters - Strategy definitions and parameters")
    print("  3. Calculation_Assumptions - All calculation assumptions")
    print("  4. Strategy_Logic_Detail - Detailed strategy logic explanation")
    print("  5. Summary_Statistics - Performance summary comparison")

if __name__ == "__main__":
    main()
