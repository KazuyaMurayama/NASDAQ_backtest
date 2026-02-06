"""
Generate Final Authoritative Results - All Strategies Comparison
最終結果ファイル生成
"""
import pandas as pd
import numpy as np
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from backtest_engine import (
    load_data, calc_dd_signal, run_backtest, calc_metrics,
    strategy_dd_vt_volspike, strategy_baseline_dd_vt
)
from test_ens2_strategies import (
    strategy_ens2_asym_slope, strategy_ens2_slope_trendtv
)


def calc_extended_metrics(nav, strat_ret, position, dates):
    """Extended metrics including median return and positive years"""
    # Base metrics
    metrics = calc_metrics(nav, strat_ret, position, dates)

    # Yearly returns for additional metrics
    nav_df = pd.DataFrame({'nav': nav.values, 'date': dates.values})
    nav_df['year'] = pd.to_datetime(nav_df['date']).dt.year
    yearly_nav = nav_df.groupby('year')['nav'].last()
    yearly_ret = yearly_nav.pct_change()
    yearly_ret.iloc[0] = yearly_nav.iloc[0] - 1

    # Median annual return
    metrics['MedianReturn'] = yearly_ret.median()

    # Positive years count
    metrics['PositiveYears'] = (yearly_ret > 0).sum()
    metrics['TotalYears'] = len(yearly_ret)

    return metrics


def main():
    print("=" * 100)
    print("FINAL RESULTS GENERATION - All Strategies Comparison")
    print("最終結果ファイル生成")
    print("=" * 100)

    # Load data
    data_path = r"C:\Users\user\Desktop\nasdaq_backtest\NASDAQ_Dairy_since1973.csv"
    df = load_data(data_path)
    print(f"Data: {df['Date'].min().date()} to {df['Date'].max().date()}")
    print(f"Period: {len(df)/252:.1f} years\n")

    close = df['Close']
    returns = close.pct_change()
    dates = df['Date']

    results = []

    def add_strategy(lev, pos, name, category, recommended=False):
        nav, strat_ret = run_backtest(close, lev)
        metrics = calc_extended_metrics(nav, strat_ret, pos, dates)
        metrics['Strategy'] = name
        metrics['Category'] = category
        metrics['Recommended'] = recommended
        results.append(metrics)
        print(f"[{category}] {name}")
        print(f"  CAGR={metrics['CAGR']*100:.2f}%, Sharpe={metrics['Sharpe']:.3f}, "
              f"MaxDD={metrics['MaxDD']*100:.2f}%, Trades={int(metrics['Trades'])}")
        return nav

    # ==========================================================================
    # All Strategies
    # ==========================================================================

    print("--- Reference (参考) ---")
    # Buy & Hold 3x
    lev_bh = pd.Series(1.0, index=close.index)
    pos_bh = pd.Series(1.0, index=close.index)
    add_strategy(lev_bh, pos_bh, 'Buy & Hold 3x', 'Reference')

    print("\n--- Baseline (ベースライン) ---")
    # DD(-18/92)+VT(25%)
    lev, pos = strategy_baseline_dd_vt(close, returns, 0.82, 0.92, 0.25, 10)
    add_strategy(lev, pos, 'DD(-18/92)+VT(25%)', 'Baseline')

    print("\n--- R4 Top Strategies (R4検証トップ) ---")
    # DD+VT+VolSpike(1.5x)
    lev, pos = strategy_dd_vt_volspike(close, returns, 0.82, 0.92, 0.25, 10, 1.5)
    add_strategy(lev, pos, 'DD+VT+VolSpike(1.5x)', 'R4-Top')

    print("\n--- Ens2 Strategies (新規Ens2戦略) ---")
    # Ens2(Asym+Slope) max_lev=1.0 [RECOMMENDED]
    lev, pos = strategy_ens2_asym_slope(close, returns, 0.82, 0.92, 0.25, 20, 5, 1.0)
    add_strategy(lev, pos, 'Ens2(Asym+Slope)', 'Ens2', recommended=True)

    # Ens2(Slope+TrendTV) max_lev=1.0
    lev, pos = strategy_ens2_slope_trendtv(close, returns, 0.82, 0.92, 20, 5, 1.0)
    add_strategy(lev, pos, 'Ens2(Slope+TrendTV)', 'Ens2')

    # ==========================================================================
    # Create Results DataFrame
    # ==========================================================================
    results_df = pd.DataFrame(results)

    # Select and order columns
    cols = ['Category', 'Strategy', 'CAGR', 'Sharpe', 'MaxDD', 'Worst5Y',
            'MedianReturn', 'PositiveYears', 'TotalYears', 'Trades', 'Recommended']
    results_df = results_df[cols]
    results_df = results_df.sort_values('Sharpe', ascending=False)

    # ==========================================================================
    # Console Output
    # ==========================================================================
    print("\n" + "=" * 100)
    print("FINAL RANKING (Sharpeソート)")
    print("=" * 100)

    print(f"\n{'Rank':<5} {'Strategy':<25} {'CAGR':>8} {'Sharpe':>8} {'MaxDD':>8} "
          f"{'Worst5Y':>8} {'Median':>8} {'+Years':>7} {'Trades':>7}")
    print("-" * 100)

    for rank, (_, row) in enumerate(results_df.iterrows(), 1):
        rec = " *" if row['Recommended'] else ""
        print(f"{rank:<5} {row['Strategy']:<25} {row['CAGR']*100:>7.2f}% {row['Sharpe']:>8.3f} "
              f"{row['MaxDD']*100:>7.2f}% {row['Worst5Y']*100:>7.2f}% "
              f"{row['MedianReturn']*100:>7.2f}% {int(row['PositiveYears']):>3}/{int(row['TotalYears'])} "
              f"{int(row['Trades']):>7}{rec}")

    print("-" * 100)
    print("* = Recommended Strategy (推奨戦略)")

    # ==========================================================================
    # Save to Excel
    # ==========================================================================
    excel_path = r"C:\Users\user\Desktop\nasdaq_backtest\FINAL_RESULTS.xlsx"

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Yu Gothic')
    normal_font = Font(size=11, name='Yu Gothic')
    bold_font = Font(bold=True, size=11, name='Yu Gothic')
    title_font = Font(bold=True, size=16, name='Yu Gothic')
    recommend_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1: Final Results
    ws1 = wb.active
    ws1.title = "最終結果"

    # Title
    ws1.merge_cells('A1:I1')
    ws1.cell(row=1, column=1, value="3x NASDAQ投資戦略 最終検証結果")
    ws1.cell(row=1, column=1).font = title_font
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:I2')
    ws1.cell(row=2, column=1, value=f"検証期間: 1974-2021 (47年間) | データ: NASDAQ Composite | max_lev=1.0 (実効レバレッジ上限3倍)")
    ws1.cell(row=2, column=1).font = Font(size=10, name='Yu Gothic', italic=True)
    ws1.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Headers
    headers = ['順位', '戦略名', 'CAGR', 'Sharpe', 'MaxDD', 'Worst5Y', '中央値\n年リターン', 'プラス\n年数', '取引\n回数']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[4].height = 35

    # Data rows
    for rank, (_, row) in enumerate(results_df.iterrows(), 1):
        row_idx = rank + 4

        ws1.cell(row=row_idx, column=1, value=rank).border = thin_border
        ws1.cell(row=row_idx, column=2, value=row['Strategy']).border = thin_border
        ws1.cell(row=row_idx, column=3, value=f"{row['CAGR']*100:.2f}%").border = thin_border
        ws1.cell(row=row_idx, column=4, value=f"{row['Sharpe']:.3f}").border = thin_border
        ws1.cell(row=row_idx, column=5, value=f"{row['MaxDD']*100:.2f}%").border = thin_border
        ws1.cell(row=row_idx, column=6, value=f"{row['Worst5Y']*100:.2f}%").border = thin_border
        ws1.cell(row=row_idx, column=7, value=f"{row['MedianReturn']*100:.2f}%").border = thin_border
        ws1.cell(row=row_idx, column=8, value=f"{int(row['PositiveYears'])}/{int(row['TotalYears'])}").border = thin_border
        ws1.cell(row=row_idx, column=9, value=int(row['Trades'])).border = thin_border

        for col in range(1, 10):
            ws1.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
            ws1.cell(row=row_idx, column=col).font = normal_font

        # Highlight recommended
        if row['Recommended']:
            for col in range(1, 10):
                ws1.cell(row=row_idx, column=col).fill = recommend_fill
                ws1.cell(row=row_idx, column=col).font = bold_font

    # Column widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 26
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 10
    ws1.column_dimensions['I'].width = 8

    # Legend
    row_idx = len(results_df) + 6
    ws1.cell(row=row_idx, column=1, value="■ 黄色ハイライト = 推奨戦略")
    ws1.cell(row=row_idx, column=1).font = Font(size=10, name='Yu Gothic')
    ws1.merge_cells(f'A{row_idx}:I{row_idx}')

    # ==========================================================================
    # Sheet 2: Strategy Details
    # ==========================================================================
    ws2 = wb.create_sheet("戦略詳細")

    ws2.merge_cells('A1:D1')
    ws2.cell(row=1, column=1, value="戦略コンポーネント詳細")
    ws2.cell(row=1, column=1).font = title_font

    details = [
        ("", "", "", ""),
        ("■ 推奨戦略: Ens2(Asym+Slope)", "", "", ""),
        ("", "", "", ""),
        ("コンポーネント", "パラメータ", "説明", ""),
        ("Layer 1: DD Control", "Exit=-18%, Reentry=92%", "200日高値比でCASH/HOLD判定", ""),
        ("Layer 2: AsymEWMA Vol", "Span: down=5, up=20", "下落時は高速反応、上昇時は低速反応", ""),
        ("Layer 3: VT Leverage", "Target Vol=25%, max_lev=1.0", "レバレッジ = min(TV/Vol, 1.0)", ""),
        ("Layer 4: SlopeMult", "MA200, Z-score 60日窓", "乗数 = clip(0.7+0.3×z, 0.3, 1.5)", ""),
        ("", "", "", ""),
        ("■ 次点戦略: DD+VT+VolSpike(1.5x)", "", "", ""),
        ("", "", "", ""),
        ("コンポーネント", "パラメータ", "説明", ""),
        ("Layer 1: DD Control", "Exit=-18%, Reentry=92%", "200日高値比でCASH/HOLD判定", ""),
        ("Layer 2: EWMA Vol", "Span=10", "標準EWMA Vol計算", ""),
        ("Layer 3: VT Leverage", "Target Vol=25%, max_lev=1.0", "レバレッジ = min(TV/Vol, 1.0)", ""),
        ("Layer 4: VolSpike", "閾値=1.5x", "Vol急騰時にレバレッジ半減", ""),
    ]

    for row_idx, (c1, c2, c3, c4) in enumerate(details, 3):
        ws2.cell(row=row_idx, column=1, value=c1)
        ws2.cell(row=row_idx, column=2, value=c2)
        ws2.cell(row=row_idx, column=3, value=c3)

        if c1.startswith("■"):
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12, name='Yu Gothic', color='4472C4')
            ws2.merge_cells(f'A{row_idx}:D{row_idx}')
        elif c1 == "コンポーネント":
            for col in range(1, 4):
                ws2.cell(row=row_idx, column=col).font = bold_font
                ws2.cell(row=row_idx, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 40

    # ==========================================================================
    # Sheet 3: Calculation Assumptions
    # ==========================================================================
    ws3 = wb.create_sheet("計算前提")

    ws3.merge_cells('A1:B1')
    ws3.cell(row=1, column=1, value="計算前提")
    ws3.cell(row=1, column=1).font = title_font

    assumptions = [
        ("データ期間", "1974年1月2日 ～ 2021年5月7日（47年間）"),
        ("データソース", "NASDAQ Composite Index (Yahoo Finance)"),
        ("ベースレバレッジ", "3倍（日次リバランス）"),
        ("max_lev", "1.0（実効レバレッジ上限 = 3倍）"),
        ("年間コスト", "0.9%（経費率、投資時のみ日割り控除）"),
        ("現金保有時リターン", "0%（金利なし）"),
        ("取引カウント", "バイナリ状態遷移のみ（HOLD↔CASH）"),
        ("", ""),
        ("指標定義", ""),
        ("CAGR", "(最終NAV)^(1/年数) - 1"),
        ("Sharpe", "(平均日次リターン×252) / (日次リターン標準偏差×√252)"),
        ("MaxDD", "ピークからの最大下落率"),
        ("Worst5Y", "ローリング5年CAGRの最小値"),
        ("中央値年リターン", "年次リターンの中央値"),
    ]

    for row_idx, (param, value) in enumerate(assumptions, 3):
        ws3.cell(row=row_idx, column=1, value=param)
        ws3.cell(row=row_idx, column=2, value=value)
        if param and not param.startswith(" "):
            ws3.cell(row=row_idx, column=1).font = bold_font

    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 55

    # ==========================================================================
    # Save
    # ==========================================================================
    wb.save(excel_path)
    print(f"\n\nExcel saved: {excel_path}")

    # ==========================================================================
    # Save CSV
    # ==========================================================================
    csv_path = r"C:\Users\user\Desktop\nasdaq_backtest\FINAL_RESULTS.csv"

    # Format for CSV
    csv_df = results_df.copy()
    csv_df['CAGR'] = csv_df['CAGR'].apply(lambda x: f"{x*100:.2f}%")
    csv_df['Sharpe'] = csv_df['Sharpe'].apply(lambda x: f"{x:.3f}")
    csv_df['MaxDD'] = csv_df['MaxDD'].apply(lambda x: f"{x*100:.2f}%")
    csv_df['Worst5Y'] = csv_df['Worst5Y'].apply(lambda x: f"{x*100:.2f}%")
    csv_df['MedianReturn'] = csv_df['MedianReturn'].apply(lambda x: f"{x*100:.2f}%")
    csv_df['PositiveYears'] = csv_df.apply(lambda r: f"{int(r['PositiveYears'])}/{int(r['TotalYears'])}", axis=1)
    csv_df['Trades'] = csv_df['Trades'].astype(int)
    csv_df = csv_df.drop(columns=['TotalYears'])

    csv_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"CSV saved: {csv_path}")

    # ==========================================================================
    # Generate Markdown
    # ==========================================================================
    md_path = r"C:\Users\user\Desktop\nasdaq_backtest\FINAL_RESULTS.md"

    md_content = """# 3x NASDAQ投資戦略 最終検証結果

## 検証概要
- **期間**: 1974年1月2日 ～ 2021年5月7日（47年間）
- **対象**: NASDAQ Composite Index
- **レバレッジ**: 3倍日次リバランス（max_lev=1.0）
- **コスト**: 年率0.9%

---

## 最終ランキング

| 順位 | 戦略名 | CAGR | Sharpe | MaxDD | Worst5Y | 中央値年リターン | プラス年数 | 取引回数 |
|------|--------|------|--------|-------|---------|-----------------|-----------|---------|
"""

    for rank, (_, row) in enumerate(results_df.iterrows(), 1):
        rec = " **[推奨]**" if row['Recommended'] else ""
        md_content += f"| {rank} | {row['Strategy']}{rec} | {row['CAGR']*100:.2f}% | {row['Sharpe']:.3f} | {row['MaxDD']*100:.2f}% | {row['Worst5Y']*100:.2f}% | {row['MedianReturn']*100:.2f}% | {int(row['PositiveYears'])}/{int(row['TotalYears'])} | {int(row['Trades'])} |\n"

    md_content += """
---

## 推奨戦略: Ens2(Asym+Slope)

### なぜこの戦略か？
1. **Sharpe 1.031** - 全戦略中最高
2. **Worst5Y +1.41%** - 唯一のプラス圏（最悪の5年間でも利益）
3. **MaxDD -48.17%** - ベースライン比で約14%改善
4. **取引回数 30回** - 低頻度で実行容易

### 戦略構成
```
Layer 1: DD Control
  - Exit: 200日高値から-18%下落 → CASH
  - Reentry: 92%回復 → HOLD

Layer 2: AsymEWMA Volatility
  - 下落時: EWMA Span=5（高速反応）
  - 上昇時: EWMA Span=20（低速反応）

Layer 3: VT Leverage
  - leverage = min(0.25 / AsymVol, 1.0)

Layer 4: SlopeMult
  - MA200の傾きをZ-score化
  - 乗数 = clip(0.7 + 0.3 × z, 0.3, 1.5)
```

### 日次判定フロー
```
1. DD判定: ratio = 終値 / 200日高値
   - ratio ≤ 0.82 → CASH
   - ratio ≥ 0.92 → HOLD

2. AsymVol計算:
   - ret < 0: α = 2/(5+1) = 0.333
   - ret ≥ 0: α = 2/(20+1) = 0.095
   - var[t] = (1-α) × var[t-1] + α × ret²
   - vol = √(var × 252)

3. VT Leverage: lev = min(0.25 / vol, 1.0)

4. SlopeMult:
   - slope = (MA200[t] - MA200[t-1]) / MA200[t-1]
   - z = (slope - mean60) / std60
   - mult = clip(0.7 + 0.3 × z, 0.3, 1.5)

5. Final: position = DD_signal × lev × mult
```

---

## 次点戦略: DD+VT+VolSpike(1.5x)

### 特徴
- **シンプル実装** - コンポーネント数が少ない
- **Sharpe 0.902** - 安定した実績
- **理解しやすい** - ロジックが直感的

### 日次判定フロー
```
1. DD判定（同上）
2. EWMA Vol (Span=10)
3. VT Leverage = min(0.25 / vol, 1.0)
4. VolSpike: 今日Vol/昨日Vol > 1.5 → lev × 0.5
5. Final: position = DD_signal × lev
```

---

## 計算前提

| 項目 | 値 |
|------|-----|
| ベースレバレッジ | 3倍（日次リバランス） |
| max_lev | 1.0（実効レバ上限3倍） |
| 年間コスト | 0.9% |
| 現金リターン | 0% |
| 取引カウント | HOLD↔CASH遷移のみ |

---

*Generated: """ + pd.Timestamp.now().strftime('%Y-%m-%d %H:%M') + "*\n"

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(md_content)

    print(f"Markdown saved: {md_path}")

    return results_df


if __name__ == "__main__":
    results = main()
